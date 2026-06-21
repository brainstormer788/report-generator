from tkinter import Tk, filedialog, messagebox
import pandas as pd
from collections import defaultdict

CATEGORIES = ["LPG", "SALP", "ALP"]
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def format_output_excel(filename):

    wb = load_workbook(filename)
    ws = wb.active

    bold_font = Font(bold=True)

    date_fill = PatternFill(
        fill_type="solid",
        start_color="D9EAD3",
        end_color="D9EAD3"
    )

    total_fill = PatternFill(
        fill_type="solid",
        start_color="FFF2CC",
        end_color="FFF2CC"
    )

    header_fill = PatternFill(
        fill_type="solid",
        start_color="DDEBF7",
        end_color="DDEBF7"
    )

    # Header row
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill

    for row in ws.iter_rows():

        first_col = str(row[0].value or "")
        second_col = str(row[1].value or "")

        # Date rows
        if first_col.startswith("Date :"):
            for cell in row:
                cell.font = bold_font
                cell.fill = date_fill

        # TOTAL rows
        if second_col == "TOTAL" or second_col == "G TOTAL":
            for cell in row:
                cell.font = bold_font
                cell.fill = total_fill

    # Auto width
    for column in ws.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 3

    wb.save(filename)

def read_file_auto(file_path):
    for header_row in range(5):
        try:
            if file_path.lower().endswith(".ods"):
                df = pd.read_excel(file_path, engine="odf", header=header_row)
            else:
                df = pd.read_excel(file_path, header=header_row)

            cols = [
                str(c).replace("\n", " ").strip().upper()
                for c in df.columns
            ]

            has_desig = "ORIGDESIG" in cols
            has_available = any("AVAILABLE" in c for c in cols)

            if has_desig and has_available:
                return df
        except Exception:
            pass

    raise ValueError(
        "Could not locate ORIGDESIG / AVAILABLE columns"
    )


def find_column(df, target):
    target = target.strip().upper()

    for col in df.columns:
        col_clean = str(col).replace("\n", " ").strip().upper()

        if target == col_clean:
            return col

    raise ValueError(
        f"Column '{target}' not found.\nColumns detected:\n{list(df.columns)}"
    )


def hour_bucket(dt):
    start_hour = dt.hour
    end_hour = (start_hour + 1) % 24
    return f"{start_hour:02d}--{end_hour:02d}"


root = Tk()
root.withdraw()

file_path = filedialog.askopenfilename(
    title="Select Crew Availability Report",
    filetypes=[
        ("Excel Files", "*.xlsx *.xls"),
        ("ODS Files", "*.ods")
    ]
)

if not file_path:
    raise SystemExit("No file selected")

try:
    df = read_file_auto(file_path)

    desig_col = find_column(df, "ORIGDESIG")

    

    available_col = next(
        c for c in df.columns
        if "AVAILABLE" in str(c).upper()
    )

    counts = defaultdict(
        lambda: {
         
            "LPG Available": 0,
           
            "SALP Available": 0,
          
            "ALP Available": 0,
        }
    )

    for _, row in df.iterrows():
        category = str(row[desig_col]).strip().upper()

        if category not in CATEGORIES:
            continue

        

        
        available_dt = pd.to_datetime(
            row[available_col],
            dayfirst=True,
            errors="coerce"
        )

        if pd.notna(available_dt):
            date_str = available_dt.strftime("%d/%m/%y")
            key = (date_str, hour_bucket(available_dt))
            counts[key][f"{category} Available"] += 1

    rows = []

    grand_totals = {
      
        "LPG Available": 0,
       
        "SALP Available": 0,
      
        "ALP Available": 0,
    }

    all_dates = sorted(
        set(date for date, _ in counts.keys()),
        key=lambda x: pd.to_datetime(x, dayfirst=True)
    )

    for date_str in all_dates:

        rows.append({
            "Date": f"Date : {date_str}",
            "Time Slot": "",
          
            "LPG Available": "",
         
            "SALP Available": "",
           
            "ALP Available": ""
        })

        date_total = {
           
            "LPG Available": 0,
        
            "SALP Available": 0,
         
            "ALP Available": 0,
        }

        for hour in range(24):
            slot = f"{hour:02d}--{(hour + 1) % 24:02d}"
            values = counts[(date_str, slot)]

            rows.append({
                "Date": "",
                "Time Slot": slot,
                **values
            })

            for col in date_total:
                date_total[col] += values[col]
                grand_totals[col] += values[col]

        rows.append({
            "Date": "",
            "Time Slot": "TOTAL",
            **date_total
        })

        rows.append({
            "Date": "",
            "Time Slot": "",
             " ": "",
            "LPG Available": "",
            " ": "",
            "SALP Available": "",
            " ": "",
            "ALP Available": ""
        })

    rows.append({
        "Date": "",
        "Time Slot": "G TOTAL",
        **grand_totals
    })


except Exception as e:
    messagebox.showerror("Error", str(e))


from datetime import datetime
from pathlib import Path


def save_report(output_df):
    """
    Save report using computer execution timestamp.
    Creates Reports folder automatically.
    Never overwrites previous reports.
    """

    Path("Reports").mkdir(exist_ok=True)

    timestamp = datetime.now().strftime(
        "%d-%m-%Y_%H-%M-%S"
    )

    output_file = (
        f"Reports/Crew_Report_{timestamp}.xlsx"
    )

    output_df.to_excel(
        output_file,
        sheet_name="Crew Summary",
        index=False
    )

    return output_file

output = pd.DataFrame(rows)
output_file = save_report(output)

messagebox.showinfo(
    "Success",
    f"Report saved successfully.\n\n{output_file}"
)
    

